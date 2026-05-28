import streamlit as st
import pandas as pd
import requests
from concurrent.futures import ThreadPoolExecutor
from io import BytesIO
import time
import random
from openpyxl.styles import PatternFill

# --- KONFIGURASI USER-AGENT ---
USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:109.0) Gecko/20100101 Firefox/119.0"
]

# --- FUNGSI INTI TRANSLATE ---
def translate_core(text, target, source='id'):
    if not text or str(text).strip().lower() in ["nan", "none", ""]:
        return ""
    
    headers = {"User-Agent": random.choice(USER_AGENTS)}
    base_url = "https://translate.googleapis.com/translate_a/single"
    
    params = {
        "client": "gtx",
        "sl": source,
        "tl": target,
        "dt": "t",
        "q": str(text).strip()
    }
    
    try:
        response = requests.get(base_url, params=params, headers=headers, timeout=12)
        if response.status_code == 200:
            result_json = response.json()
            translated_parts = [part[0] for part in result_json[0] if part[0]]
            return "".join(translated_parts)
        elif response.status_code == 429:
            return "ERR_LIMIT"
    except Exception:
        pass
    return None

def translate_smart(text, target):
    text_str = str(text).strip()
    if len(text_str) <= 4500:
        return translate_core(text_str, target)
    
    chunks = [text_str[i:i+4000] for i in range(0, len(text_str), 4000)]
    translated_results = []
    for c in chunks:
        res = translate_core(c, target)
        if res and res != "ERR_LIMIT":
            translated_results.append(res)
        else:
            return "ERR_LIMIT" if res == "ERR_LIMIT" else None
    return " ".join(translated_results)

# --- ANTARMUKA STREAMLIT ---
st.set_page_config(page_title="Turbo Translator Pro v2 (Multi-Sheet)", page_icon="⚡", layout="wide")
st.title("⚡ Turbo Excel Translator (Semua Sheet)")
st.markdown("Alat translasi otomatis untuk **semua sheet** file Excel. Kolom B (indeks 1) di setiap sheet akan diterjemahkan.")
st.markdown("Gunakan 1 tab saja, dan pastikan setiap sheet memiliki kolom B yang ingin diterjemahkan.")

# --- SIDEBAR ---
st.sidebar.header("⚙️ Pengaturan")
target_lang = st.sidebar.text_input("Kode Bahasa Tujuan", value="en", help="Contoh: en, ja, fi, ko, ar")
max_workers = st.sidebar.slider("Kecepatan (Workers)", 1, 15, 5, help="Disarankan 5-10 agar aman.")
st.sidebar.markdown("---")
st.sidebar.info("📌 Jika hasil unduhan berwarna merah, IP-mu terkena limit. Kurangi Workers atau ganti koneksi.")

# --- PROSES UTAMA ---
uploaded_file = st.file_uploader("Upload file Excel (.xlsx)", type=["xlsx"])

if uploaded_file:
    try:
        # BACA SEMUA SHEET
        all_sheets = pd.read_excel(uploaded_file, sheet_name=None)  # dict: {nama_sheet: DataFrame}
        sheet_names = list(all_sheets.keys())
        total_sheets = len(sheet_names)

        st.success(f"File '{uploaded_file.name}' berhasil dimuat. Ditemukan {total_sheets} sheet: {', '.join(sheet_names)}")

        if st.button("🚀 Mulai Terjemahkan Semua Sheet"):
            # Persiapan data: kumpulkan semua teks dari setiap sheet (kolom B)
            tasks = []  # (sheet_name, row_index, text)
            for sheet_name, df in all_sheets.items():
                if df.shape[1] < 2:
                    st.warning(f"Sheet '{sheet_name}' tidak memiliki kolom B, dilewati.")
                    continue
                col_b = df.iloc[:, 1].astype(str).tolist()
                for row_idx, text in enumerate(col_b):
                    tasks.append((sheet_name, row_idx, text))

            total_tasks = len(tasks)
            if total_tasks == 0:
                st.error("Tidak ada teks yang bisa diterjemahkan.")
            else:
                # Proses paralel
                results = [None] * total_tasks
                progress_bar = st.progress(0)
                status_placeholder = st.empty()
                time_placeholder = st.empty()
                start_time = time.time()

                # Fungsi pembantu untuk eksekusi paralel
                def process_task(task):
                    sheet_name, row_idx, text = task
                    return translate_smart(text, target_lang)

                with ThreadPoolExecutor(max_workers=max_workers) as executor:
                    future_to_idx = {executor.submit(process_task, tasks[i]): i for i in range(total_tasks)}
                    completed = 0
                    for future in future_to_idx:
                        idx = future_to_idx[future]
                        try:
                            results[idx] = future.result()
                        except:
                            results[idx] = None
                        completed += 1

                        elapsed = time.time() - start_time
                        avg_time = elapsed / completed if completed > 0 else 0
                        eta = int(avg_time * (total_tasks - completed))
                        progress_bar.progress(completed / total_tasks)
                        status_placeholder.write(f"⏳ Memproses: {completed}/{total_tasks} teks dari semua sheet")
                        time_placeholder.markdown(f"⏱️ Sisa waktu: **{eta} detik**")

                # Memasukkan hasil kembali ke DataFrame masing-masing
                # Buat mapping (sheet_name, row_idx) -> hasil
                result_map = {}
                for i, (sheet_name, row_idx, _) in enumerate(tasks):
                    result_map[(sheet_name, row_idx)] = results[i]

                # Update setiap DataFrame
                for sheet_name, df in all_sheets.items():
                    if df.shape[1] >= 2:
                        hasil_kolom = []
                        for row_idx in range(len(df)):
                            hasil_kolom.append(result_map.get((sheet_name, row_idx), ""))
                        df['Hasil Translate'] = hasil_kolom

                # --- PREVIEW (Sheet pertama yang memiliki kolom B) ---
                first_valid_sheet = None
                for sn in sheet_names:
                    if all_sheets[sn].shape[1] >= 2:
                        first_valid_sheet = sn
                        break
                if first_valid_sheet:
                    st.subheader(f"📋 Preview Sheet '{first_valid_sheet}' (5 Baris Pertama)")
                    st.dataframe(all_sheets[first_valid_sheet][['Hasil Translate']].head(5))

                # --- GENERASI FILE OUTPUT DENGAN WARNA ---
                nama_file_murni = uploaded_file.name.rsplit('.', 1)[0]
                nama_file_baru = f"{nama_file_murni} ({target_lang}).xlsx"
                output = BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    for sheet_name, df in all_sheets.items():
                        df.to_excel(writer, index=False, sheet_name=sheet_name[:31])  # max 31 karakter
                        worksheet = writer.sheets[sheet_name[:31]]
                        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        # Cek hasil terjemahan, warnai merah jika error/kosong
                        if 'Hasil Translate' in df.columns:
                            hasil = df['Hasil Translate'].tolist()
                            for row_num, val in enumerate(hasil, start=2):  # start=2 karena header baris 1
                                if val is None or val == "ERR_LIMIT" or val == "":
                                    for col_num in range(1, df.shape[1] + 1):
                                        worksheet.cell(row=row_num, column=col_num).fill = red_fill

                st.success(f"✅ Selesai! Semua sheet diterjemahkan. File: {nama_file_baru}")
                st.download_button(
                    label="📥 Download Hasil Terjemahan (Semua Sheet)",
                    data=output.getvalue(),
                    file_name=nama_file_baru,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

    except Exception as e:
        st.error(f"Terjadi kesalahan: {e}")
